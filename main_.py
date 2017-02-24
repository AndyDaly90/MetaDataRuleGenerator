#!/usr/bin/python
# -*- coding: utf-8 -*-

from Tkinter import *
import Tkinter as tk
from tkinter import filedialog
from metaDataRules import *
import tkinter.messagebox as message
import openpyxl
import re


def display_setValue():
    result = MetaDataRules.setValue(form_id.get(), destination_field_id.get(), set_value.get())
    entry_box.insert(1.0, result)


def display_setConcat():
    result = MetaDataRules.concat(form_id.get(), destination_field_id.get(), first_field_id.get(), concat_val.get(),
                                  second_field_id.get())
    entry_box.insert(1.0, result)


def display_delete():
    result = MetaDataRules.delete(form_id.get(), delete_field.get())
    entry_box.insert(1.0, result)


def display_setDateUDT():
    result = MetaDataRules.stringToDate(d_field.get(), s_field.get())
    entry_box.insert(1.0, result)


def displayBI():
    result = MetaDataRules.validate_BIC_IBAN(bic_or_iban.get(), source_id.get(), destination_id.get())
    entry_box.insert(1.0, result)


def display_crossDB_insert():
    result = MetaDataRules.crossDbInsert(source_form_id, source_form, source_field_IDs,
                                         destination_form_id, destination_fieldIDs)
    entry_box.insert(1.0, result)


def display_crossDB_retrieval():
    result = MetaDataRules.crossDbRetrieval(destination_fieldIDs, source_form_id, source_field_IDs,
                                            source_ids_toMatch, destination_values_toMatch)
    entry_box.insert(1.0, result)


def display_custom_file_output():
    result = MetaDataRules.customFileOutput(source_headers, source_fieldIDs, source_id, export_title.get(),
                                            file_path.get(), variable.get())
    entry_box.delete(1.0, 'end')
    entry_box.insert(1.0, result)


def copy_to_clipboard():
    clip = Tk()
    clip.withdraw()
    clip.clipboard_clear()
    clip.clipboard_append(entry_box.get(1.0, END))
    clip.destroy()

class MainWindow(tk.Frame):
    def __init__(self, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)
        menu = tk.Menu(root)
        root.config(menu=menu)

        sub_process_string = tk.Menu(menu)
        sub_process_date = tk.Menu(menu)
        sub_process_bic_iban = tk.Menu(menu)
        sub_process_cross_db = tk.Menu(menu)
        sub_process_import_export = tk.Menu(menu)

        menu.add_cascade(label="String Process Rules", menu=sub_process_string)
        menu.add_cascade(label="Date Process Rules", menu=sub_process_date)
        menu.add_cascade(label="Validation", menu=sub_process_bic_iban)
        menu.add_cascade(label="Cross Database Rules", menu=sub_process_cross_db)
        menu.add_cascade(label="Import/Export Rules", menu=sub_process_import_export)

        sub_process_string.add_command(label="Set Value", command=self.captureSetValue)
        sub_process_string.add_command(label="Concat", command=self.captureSetConcat)
        sub_process_string.add_command(label="Delete", command=self.captureDelete)
        sub_process_string.add_separator()
        # sub_process_string.add_command(label="Exit", command=leave)

        sub_process_date.add_command(label="String to Date (UDT)", command=self.captureDate)

        sub_process_bic_iban.add_command(label="Validate IBAN|BIC", command=self.captureBI)

        sub_process_cross_db.add_command(label="Cross Database Insert", command=self.captureCrossDBInsert)
        sub_process_cross_db.add_command(label="Cross Database Retrieval", command=self.captureCrossDBRetrieval)

        sub_process_import_export.add_command(label="Custom File Output TXT|CSV",
                                              command=self.captureCustomFileOutput_txt)

    def captureSetValue(self):
        window = tk.Toplevel(self)
        window.wm_title("Set Value")

        Label(window, text="Enter Form ID: ").grid(row=0)
        Label(window, text="Enter Destination Field ID: ").grid(row=1)
        Label(window, text="Enter Value to set: ").grid(row=2)

        global form_id
        global destination_field_id
        global set_value
        global entry_box

        form_id = Entry(window)

        destination_field_id = Entry(window)

        set_value = Entry(window)

        form_id.grid(row=0, column=1)

        destination_field_id.grid(row=1, column=1)

        set_value.grid(row=2, column=1)

        btn = Button(window, text="Create Rule", command=display_setValue)
        btn.grid(row=3, column=0)

        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)
        btn2.grid(row=6, column=1)

        entry_box = Text(window, width=60, height=6)
        entry_box.grid(row=4, column=1)

    def captureSetConcat(self):
        window = tk.Toplevel(self)
        window.wm_title("Concatenation")

        Label(window, text="Enter Form ID: ").grid(row=0)
        Label(window, text="Enter Destination Field ID: ").grid(row=1)
        Label(window, text="Enter First Field ID: ").grid(row=2)
        Label(window, text="Enter Concatenation Value: ").grid(row=3)

        global form_id
        global destination_field_id
        global first_field_id
        global concat_val
        global second_field_id

        form_id = Entry(window)
        destination_field_id = Entry(window)
        first_field_id = Entry(window)
        concat_val = Entry(window)
        second_field_id = Entry(window)

        form_id.grid(row=0, column=1)

        destination_field_id.grid(row=1, column=1)

        first_field_id.grid(row=2, column=1)

        concat_val.grid(row=3, column=1)

        second_field_id.grid(row=4, column=1)

        btn = Button(window, text="Create Rule", command=display_setConcat)
        btn.grid(row=4, column=0)

        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)
        btn2.grid(row=6, column=1)
        global entry_box
        entry_box = Text(window, width=60, height=6)
        entry_box.grid(row=4, column=1)

    def captureDelete(self):
        window = tk.Toplevel(self)
        window.wm_title("Set Value")

        Label(window, text="Enter Form ID: ").grid(row=0)
        Label(window, text="Enter Field ID To Delete: ").grid(row=1)

        global form_id
        global delete_field
        global entry_box

        form_id = Entry(window)
        delete_field = Entry(window)

        form_id.grid(row=0, column=1)
        delete_field.grid(row=1, column=1)

        btn = Button(window, text="Create Rule", command=display_delete)
        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)

        btn.grid(row=3, column=0)
        btn2.grid(row=4, column=1)

        entry_box = Text(window, width=60, height=4)
        entry_box.grid(row=3, column=1)

    def captureDate(self):
        window = tk.Toplevel(self)
        window.wm_title("Convert to UDT")

        Label(window, text="Enter Source Field ID: ").grid(row=0)
        Label(window, text="Enter Destination Field ID: ").grid(row=1)

        global d_field
        global s_field
        global entry_box

        s_field = Entry(window)
        d_field = Entry(window)

        s_field.grid(row=0, column=1)
        d_field.grid(row=1, column=1)

        btn = Button(window, text="Create Rule", command=display_setDateUDT)
        btn.grid(row=3, column=0)

        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)
        btn2.grid(row=6, column=1)

        entry_box = Text(window, width=60, height=6)
        entry_box.grid(row=3, column=1)

    def captureBI(self):
        window = tk.Toplevel(self)
        window.wm_title("Validate IBAN|BIC")

        Label(window, text="Enter The Source Field ID: ").grid(row=1)
        Label(window, text="Enter The Destination Field ID: ").grid(row=2)

        global source_id
        global destination_id
        global entry_box
        global bic_or_iban

        source_id = Entry(window)
        destination_id = Entry(window)

        choices = ['BIC', 'IBAN']
        bic_or_iban = StringVar(window)
        bic_or_iban.set(choices[0])

        op_menu = OptionMenu(window, bic_or_iban, *choices)

        source_id.grid(row=0, column=1)
        destination_id.grid(row=1, column=1)
        op_menu.grid(row=2, column=1)

        btn = Button(window, text="Create Rule", command=displayBI)

        entry_box = Text(window, width=60, height=6)
        entry_box.grid(row=3, column=1)

        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)
        btn.grid(row=3, column=0)
        btn2.grid(row=5, column=1)

    def captureCrossDBInsert(self):
        try:
            global window
            window = tk.Toplevel(self)
            window.wm_title("Cross Database Insert")
            message.showinfo("Process", "Please select the excel file containing the source & destination field IDs")
            root.fileName = filedialog.askopenfilename(filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))

            # This workbook object represents the excel file
            work_book = openpyxl.load_workbook(root.fileName)

            sheet1 = work_book.get_sheet_by_name('Sheet1')

            global source_field_IDs
            global destination_fieldIDs
            global source_form_id
            global destination_form_id
            global source_form
            global entry_box

            source_field_IDs = []
            destination_fieldIDs = []

            for i in range(2, sheet1.max_row + 1):
                source_field_IDs.append(sheet1.cell(row=i, column=1).value)
                destination_fieldIDs.append(sheet1.cell(row=i, column=2).value)

            source_form_id = re.findall(r'\d+', source_field_IDs[1])[:1]
            destination_form_id = re.findall(r'\d+', destination_fieldIDs[1])[:1]

            source_form_id = source_form_id[0]

            destination_form_id = destination_form_id[0]

            source_form = "strCDA_" + source_form_id + "_"

            entry_box = Text(window, width=75, height=22)
            entry_box.grid(row=3, column=1)

            btn = Button(window, text="Create Rule", command=
            display_crossDB_insert)

            btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)

            btn.grid(row=3, column=0)
            btn2.grid(row=4, column=1)

            message.showinfo("Success", "The excel file has been loaded \n Select Create Rule")
            window.geometry('730x400+300+330')

        except EnvironmentError:
            message.showerror("Result", "There is an error with your file")

    def captureCrossDBRetrieval(self):
        window = tk.Toplevel(self)
        window.wm_title("Cross Database Insert")
        message.showinfo("Process", "Please select the excel file containing the source & destination field IDs"
                         + "\nAlong with fields to match on")
        root.fileName = filedialog.askopenfilename(filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))

        # This workbook object represents the excel file
        work_book = openpyxl.load_workbook(root.fileName)

        sheet1 = work_book.get_sheet_by_name('Sheet1')

        global destination_fieldIDs
        global source_form_id
        global source_field_IDs
        global source_ids_toMatch
        global destination_values_toMatch
        global entry_box

        destination_fieldIDs = []
        source_field_IDs = []
        source_ids_toMatch = []
        destination_values_toMatch = []

        for i in range(2, sheet1.max_row + 1):
            destination_fieldIDs.append(sheet1.cell(row=i, column=1).value)
            source_field_IDs.append(sheet1.cell(row=i, column=2).value)
            source_ids_toMatch.append(sheet1.cell(row=i, column=3).value)
            destination_values_toMatch.append(sheet1.cell(row=i, column=4).value)

        print(destination_fieldIDs)

        source_form_id = re.findall(r'\d+', source_field_IDs[1])[:1]

        entry_box = Text(window, width=75, height=22)
        entry_box.grid(row=3, column=1)

        btn = Button(window, text="Create Rule", command=
        display_crossDB_retrieval)

        btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)

        btn.grid(row=3, column=0)
        btn2.grid(row=4, column=1)

        message.showinfo("Success", "The excel file has been loaded \n Select Create Rule")
        window.geometry('730x400+300+330')

    def captureCustomFileOutput_txt(self):
        try:
            window = tk.Toplevel(self)
            window.wm_title("Cross Database Insert")

            message.showinfo("Process", "Please select the excel file containing the Headers & field IDs")
            root.fileName = filedialog.askopenfilename(filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))

            # This workbook object represents the excel file
            work_book = openpyxl.load_workbook(root.fileName)

            sheet1 = work_book.get_sheet_by_name('Sheet1')

            global source_headers
            global source_fieldIDs
            global source_id
            global entry_box

            source_headers = []
            source_fieldIDs = []

            for i in range(2, sheet1.max_row + 1):
                source_headers.append(sheet1.cell(row=i, column=1).value)
                source_fieldIDs.append(sheet1.cell(row=i, column=2).value)

            source_id = re.findall(r'\d+', source_fieldIDs[1])[:1]
            source_id = source_id[0]

            entry_box = Text(window, width=75, height=22)
            entry_box.grid(row=5, column=1)

            global variable

            choices = ['TXT', 'CSV']
            variable = StringVar(window)
            variable.set(choices[0])

            op_menu = OptionMenu(window, variable, *choices)

            Label(window, text="Enter The Export Title: ").grid(row=1)
            Label(window, text="Enter The File Path: ").grid(row=2)
            Label(window, text="Select the Output format:").grid(row=3)

            global export_title
            global file_path

            export_title = Entry(window)
            file_path = Entry(window)

            export_title.grid(row=1, column=1)
            file_path.grid(row=2, column=1)
            op_menu.grid(row=3, column=1)

            btn = Button(window, text="Create Rule", command=
            display_custom_file_output)

            btn2 = Button(window, text="Copy to clipboard", command=copy_to_clipboard)

            btn.grid(row=5, column=0)
            btn2.grid(row=6, column=1)

            message.showinfo("Success", "The excel file has been loaded \n Select Create Rule")
            window.geometry('840x470+300+330')

        except EnvironmentError:
            message.showerror("Result", "There is an error with your file")


if __name__ == "__main__":
    root = tk.Tk()
    main = MainWindow(root)
    main.pack(side="top", fill="both", expand=True)
    root.geometry("900x100")
    root.mainloop()
